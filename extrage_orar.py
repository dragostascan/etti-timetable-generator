import pandas as pd
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==========================================
# 1. CONFIGURARE GLOBALA
# ==========================================
# Aici definim dicționarele de traducere. Secretariatul scrie în "Anunțuri" numele lung (ex: "Fizică"),
# dar în grila de "Orar" scrie prescurtat ("fiz2 (l)"). Aceste dicționare fac legătura între ele.

TRADUCATOR_RO = {
    # ================= ANUL 1 =================
    "fiz2 (l)": "Fizică",
    "fiz2(l)": "Fizică",
    "pclp2 (l)": "Programarea Calculatoarelor și Limbaje de Programare 2 (PCLP)",
    "pclp2(l)": "Programarea Calculatoarelor și Limbaje de Programare 2 (PCLP)",
    "metc (l)": "Măsurări în Electronică și Telecomunicații (METc)",
    "metc(l)": "Măsurări în Electronică și Telecomunicații (METc)",
    "ia (p)": "Informatica aplicata - proiect",
    "ia(p)": "Informatica aplicata - proiect",
    "so1 (l)": "Sisteme de Operare 1 (SO1)",
    "so1(l)": "Sisteme de Operare 1 (SO1)",
    "me (l)": "Materiale pentru Electronică (ME)",
    "me(l)": "Materiale pentru Electronică (ME)",
    "sport (s)": "Educaţie fizică şi sport 2",
    "sport(s)": "Educaţie fizică şi sport 2",
    "sport": "Educaţie fizică şi sport 2",

    # ================= ANUL 2 =================
    "amp (l)": "Arhitectura microprocesoarelor 2. Microcontrolere (laborator)",
    "amp(l)": "Arhitectura microprocesoarelor 2. Microcontrolere (laborator)",
    "bd (l)": "Baze de date (laborator)",
    "bd(l)": "Baze de date (laborator)",
    "cef (l)": "Circuite Electronice Fundamentale + DEEA2 - Laborator",
    "cef(l)": "Circuite Electronice Fundamentale + DEEA2 - Laborator",
    "cid (l)": "Circuite Integrate Digitale + ED - Laborator",
    "cid(l)": "Circuite Integrate Digitale + ED - Laborator",
    "ss (l)": "Semnale și sisteme 2 (laborator)",
    "ss(l)": "Semnale și sisteme 2 (laborator)",
    "ss2 (l)": "Semnale și sisteme 2 (laborator)",
    "ss2(l)": "Semnale și sisteme 2 (laborator)",
    "sp (s)": "Sport (SP) (seminar)",
    "sp(s)": "Sport (SP) (seminar)",
    "ma2 (l)": "Matematici Aplicate 2 (laborator)",
    "ma2(l)": "Matematici Aplicate 2 (laborator)",

    # ================= ANUL 3 =================
    "depi (l)": "Decizie si estimare in prelucrarea informatiilor",
    "depi(l)": "Decizie si estimare in prelucrarea informatiilor",
    "depi (s)": "Decizie si estimare in prelucrarea informatiilor",
    "depi(s)": "Decizie si estimare in prelucrarea informatiilor",
    "pds (l)": "Prelucrarea digitală a semnalelor",
    "pds(l)": "Prelucrarea digitală a semnalelor",
    "pds (s)": "Prelucrarea digitală a semnalelor",
    "pds(s)": "Prelucrarea digitală a semnalelor",
    "tv (l)": "Televiziune",
    "tv(l)": "Televiziune",
    "tpi (l)": "Tehnologia si prelucrarea informatiei",
    "tpi(l)": "Tehnologia si prelucrarea informatiei",
    "bsad (l)": "Baze de date pentru stocarea si analiza datelor",
    "bsad(l)": "Baze de date pentru stocarea si analiza datelor",
    "mic (l)": "Microprocesoare (Microprocessors)",
    "mic(l)": "Microprocesoare (Microprocessors)",
    "pr apm": "Aplicatii practice ale microcontrolerelor - Proiect",

    # ================= ANUL 4 =================
    "cm (l)": "COMUNICAŢII MOBILE",
    "cm(l)": "COMUNICAŢII MOBILE",
    "tcsm (l)": "TEHNICI DE COMPRESIE A SEMNALELOR MULTIMEDIA",
    "tcsm(l)": "TEHNICI DE COMPRESIE A SEMNALELOR MULTIMEDIA",
    "caf (l)": "CALITATE ŞI FIABILITATE",
    "caf(l)": "CALITATE ŞI FIABILITATE",
    "rcm (l)": "REŢELE DE COMUNICAŢII MOBILE",
    "rcm(l)": "REŢELE DE COMUNICAŢII MOBILE",
    "tstm (l)": "TEHNICI ŞI SISTEME DE TRANSMISIUNI MULTIPLEX",
    "tstm(l)": "TEHNICI ŞI SISTEME DE TRANSMISIUNI MULTIPLEX",
    "so (l)": "Sisteme de operare",
    "so(l)": "Sisteme de operare",
    "ici (l)": "Inteligenta computationala integrata",
    "ici(l)": "Inteligenta computationala integrata",
    "robo (l)": "Robotica",
    "robo(l)": "Robotica",
    "src (l)": "Sisteme reconfigurabile de calcul",
    "src(l)": "Sisteme reconfigurabile de calcul",
    "aaccep (l)": "Analiza asist. de calculator a circ. electr. de putere",
    "aaccep(l)": "Analiza asist. de calculator a circ. electr. de putere",
    "taep (l)": "Testarea automata a echipamentelor si proceselor",
    "taep(l)": "Testarea automata a echipamentelor si proceselor",
    "scm (l)": "SISTEME DE COMUNICAȚII MOBILE",
    "scm(l)": "SISTEME DE COMUNICAȚII MOBILE",
    "AI (l)": "Analiza imaginilor",
    "AI(l)": "Analiza imaginilor",
    "APD(l)": "Algoritmi paraleli si distribuiti 2",
    "APD (l)": "Algoritmi paraleli si distribuiti 2",
    "DDM(l)": "Dispozitive dielectrice si magnetice",
    "DDM (l)": "Dispozitive dielectrice si magnetice"
}

TRADUCATOR_EN = {
    # ================= ANUL 1 =================
    # Prescurtarile din orar sunt in engleza, dar in Anunturi scrie in romana
    "phyiscs2 (l)": "Fizică",
    "physics2(l)": "Fizică",
    "cppl2 (l)": "Programarea Calculatoarelor și Limbaje de Programare 2 (PCLP)",
    "cppl2(l)": "Programarea Calculatoarelor și Limbaje de Programare 2 (PCLP)",
    "metc (l)": "Măsurări în Electronică și Telecomunicații (METc)",
    "metc(l)": "Măsurări în Electronică și Telecomunicații (METc)",
    "ai (p)": "Informatica aplicata - proiect",
    "ai(p)": "Informatica aplicata - proiect",
    "em (l)": "Materiale pentru Electronică (ME)",
    "em(l)": "Materiale pentru Electronică (ME)",
    "sport": "Educaţie fizică şi sport 2",

    # ================= ANUL 2 =================
    # Aici anunturile incep sa aiba denumirile in engleza, iar orarul are prescurtari EN (ex: FEC)
    "ma2 (l)": "Microprocessor Architecture 2 (lab)",
    "ma2(l)": "Microprocessor Architecture 2 (lab)",
    "db (l)": "Data bases (laborator)",
    "db(l)": "Data bases (laborator)",
    "fec (l)": "Fundamental Electronic Circuits - Laboratory",
    "fec(l)": "Fundamental Electronic Circuits - Laboratory",
    "dic (l)": "Digital Integrated Circuits - Laboratory",
    "dic(l)": "Digital Integrated Circuits - Laboratory",
    "ss (l)": "Signals and systems 2 (lab)",
    "ss(l)": "Signals and systems 2 (lab)",
    "sport": "Sport (SP) (seminar)",

    # ================= ANUL 3 =================
    "deip (l)": "Decizie si estimare in prelucrarea informatiilor",
    "deip(l)": "Decizie si estimare in prelucrarea informatiilor",
    "dsp (l) SC2": "Prelucrarea digitală a semnalelor",
    "dsp(l) SC2": "Prelucrarea digitală a semnalelor",
    "mc(l)": "Circuite de microunde",
    "mc (l)": "Circuite de microunde",
    "tv (l)": "Televiziune",
    "tv(l)": "Televiziune",
    "nai (l) Sc.2": "Network Architectures and Internet",
    "bsad(l)": "Bazele sistemelor de achiziție de date",
    "bsad (l)": "Bazele sistemelor de achiziție de date",
    "tdav(l)": "Tehnici digitale audio video",
    "tdav (l)": "Tehnici digitale audio video",
    
    # ================= ANUL 4 =================
    "robo (l)": "Robotica",
    "robo(l)": "Robotica",
    "mcs (l)": "Mobile communications systems",
    "mcs(l)": "Mobile cummunications systems",
    "CAAPEC (l)": "Analiza asist. de calculator a circ. electr. de putere",
    "CAAPEC(l)": "Analiza asist. de calculator a circ. electr. de putere",
    "CAF(l)": "CALITATE ŞI FIABILITATE",
    "CAF (l)": "CALITATE ŞI FIABILITATE",
    "MC (l)": "MOBILE COMMUNICATIONS",
    "MC(l)": "MOBILE COMMUNICATIONS",
    "MC-TA (l)": "MULTIMEDIA CODING - TECHNIQUES AND APPLICATIONS",
    "MC-TA(l)": "MULTIMEDIA CODING - TECHNIQUES AND APPLICATIONS",
    "ATEP (l)": "Automatic testing of equipments and processes",
    "ATEP(l)": "Automatic testing of equipments and processes",
    "rcs (l)": "Sisteme reconfigurabile de calcul",
    "rcs(l)": "Sisteme reconfigurabile de calcul",
    "QR (l) SC1": "Quality and Reliability",
    "QR(l) SC1": "Quality and Reliability"
}

ZILE_SAPTAMANA = ['LUNI', 'MARTI', 'MIERCURI', 'JOI', 'VINERI']
# Scurtatura pt a genera lista de ore: ['09-10', '10-11', ... , '20-21']
ORE_STANDARD = [f"{i:02d}-{i+1:02d}" for i in range(9, 21)]

# ==========================================
# 2. FUNCTII UTILE (Curatare si Procesare logica)
# ==========================================
def extrage_grupa_si_anul():
    # Verifica daca grupa respecta formatul facultatii (ex: 432A -> Anul 3, Seria A, Grupa 2)
    grupa = input("Te rog introdu grupa ta (ex: 422A, 414C, 411G): ").strip().upper()
    if len(grupa) == 4 and grupa[:3].isdigit() and grupa[3].isalpha() and 1 <= int(grupa[1]) <= 4:
        return grupa, grupa[1] 
    else:
        print("Eroare: Formatul grupei este invalid! Te rog sa scrii o grupa reala (ex: 411C, 432A).")
        return extrage_grupa_si_anul()

def este_materie(text):
    # Opreste anumite formatari din Excel (cratime, spatii goale, "nan")
    t = str(text).strip()
    if re.match(r'^-+$', t): # Daca are doar cratime (---)
        return False
    return t not in ["", "Liber", "nan"]

def normalize_text(text):
    # Doar taie spatiile invizibile de la margini
    return str(text).strip()

def normalize_key(text):
    # Face toate literele mici, sterge diacriticele si spatiile duble, face legatura cu Traducatorul
    text = str(text).strip().lower()
    replacements = {
        'ă': 'a', 'â': 'a', 'î': 'i', 'ș': 's', 'ş': 's', 'ț': 't', 'ţ': 't'
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = re.sub(r'\s+', ' ', text)
    return text

def scoate_sala_din_text(text):
    # Extrage materia pura. Ex: "Fizica [A05]" -> "Fizica"
    text = str(text).strip()
    text = re.sub(r'\s*\[[^\]]+\]\s*', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def extrage_sala_din_text(text):
    # Extrage DOAR sala. Ex: "Fizica [A05]" -> "A05"
    text = str(text).strip()
    m = re.search(r'\[([^\]]+)\]', text)
    return m.group(1).strip() if m else ""

def inlocuieste_sala_in_text(text, sala_noua):
    # Sterge sala veche (daca exista) si o lipeste pe cea noua la final
    baza = scoate_sala_din_text(text)
    if sala_noua:
        return f"{baza} [{sala_noua}]"
    return baza

def materia_are_sala(text):
    # Verifica daca materia are deja sala atasata in vreun fel (in paranteze sau pur si simplu codul langa)
    if not text or str(text).strip() == "Liber": return False
    text = str(text)
    if re.search(r'\[[^\]]+\]', text): return True
    if re.search(r'\b[A-Za-z]{0,3}\d{2,4}[A-Za-z]?\b', text): return True
    return False

def parseaza_text_materie(text):
    # Daca vede "Fizica / Sport", stie ca Fizica e la Sapt Impara, iar Sport la Para.
    text = str(text)
    if text == 'nan' or text.strip() == '':
        return {"Impara": "Liber", "Para": "Liber"}

    if text.strip() == "Liber" or re.match(r'^-+$', text.strip()):
        return {"Impara": "Liber", "Para": "Liber"}

    linii = text.split('\n') # Taie randurile separate cu Alt+Enter
    impara_final = []
    para_final = []

    for linie in linii:
        linie = linie.strip()
        if not linie: continue

        if '/' in linie: # Daca avem materii care alterneaza
            parti = linie.split('/')
            imp = parti[0].strip() # Stanga slash-ului merge in Impara
            par = parti[1].strip() if len(parti) > 1 else "" # Dreapta merge in Para

            if imp and not re.match(r'^-+$', imp): impara_final.append(imp)
            if par and not re.match(r'^-+$', par): para_final.append(par)
        else: # Daca e materie normala saptamanala, o punem in ambele
            if not re.match(r'^-+$', linie):
                impara_final.append(linie)
                para_final.append(linie)

    val_imp = " ".join(impara_final).strip()
    val_par = " ".join(para_final).strip()

    # Returneaza un "pachet" (dictionar) cu ce facem in fiecare saptamana
    return {
        "Impara": val_imp if val_imp else "Liber",
        "Para": val_par if val_par else "Liber"
    }

def extrage_ore_slots(ora_str):
    # Transforma "10-12" din Excel intr-o lista de ore individuale: ["10-11", "11-12"]
    ora_str = str(ora_str).strip()
    # Stergem detaliile de genul "sapt 1-7" ca sa ramanem doar cu ora
    ora_str = re.sub(r'(?i)s[aă]pt[a-z\.\s]*\d{1,2}\s*-\s*\d{1,2}', '', ora_str)

    # Bug comun din Excel: Uneori "09-11" e transformat automat in data "2016-09-11". Aici il detectam.
    m_date = re.search(r'\d{4}-(\d{2})-(\d{2})', ora_str)
    if m_date:
        h1, h2 = int(m_date.group(1)), int(m_date.group(2))
        if 9 <= h1 <= 20 and 10 <= h2 <= 21 and h1 < h2:
            return [f"{h:02d}-{h+1:02d}" for h in range(h1, h2)]

    # Cautare normala pentru tiparul "OraInceput - OraSfarsit" (ex: 14-16)
    m_norm = re.search(r'([012]?\d)\s*-\s*([012]?\d)', ora_str)
    if m_norm:
        h1, h2 = int(m_norm.group(1)), int(m_norm.group(2))
        # Validam ca orele sa fie intre limitele normale (9:00 -> 21:00) ca sa evitam erori
        if 9 <= h1 <= 20 and 10 <= h2 <= 21 and h1 < h2:
            return [f"{h:02d}-{h+1:02d}" for h in range(h1, h2)]

    return []

def este_curs(text):
    t = str(text).lower()
    return "(curs)" in t or "(c)" in t

def este_activitate_sport(text):
    t = scoate_sala_din_text(text).lower().strip()
    t = re.sub(r'\s+', ' ', t)

    variante_sport = {
        "sport",
        "sport (s)",
        "sp(s)",
        "sp (s)",
        "educatie fizica",
        "educaţie fizică",
        "educatie fizica si sport 2",
        "educaţie fizică şi sport 2"
    }

    return t in variante_sport

def sala_de_sport(sala_text):
    s = str(sala_text).lower()
    return "sport" in s or "sala sport" in s

def seria_se_potriveste(serii_text, grupa_cautata):
    # Citeste textul "A, C, G" din foaia de anunturi si verifica daca apartii listei
    serii_text = normalize_text(serii_text).upper()
    grupa_cautata = str(grupa_cautata).strip().upper()
    
    # Extrage doar litera seriei
    seria_mea = ""
    for char in reversed(grupa_cautata):
        if char.isalpha():
            seria_mea = char
            break
            
    if serii_text in ["TOATE", "TOTI", "TOȚI", "ALL"]: return True

    if '-' in serii_text: # Daca e interval (ex: A-F)
        capete = serii_text.split('-')
        if len(capete) == 2:
            st = capete[0].strip().upper()
            dr = capete[1].strip().upper()
            if len(st) == 1 and len(dr) == 1 and len(seria_mea) == 1:
                if st <= seria_mea <= dr:
                    return True

    # Desparte lista prin virgula si cauta seria ta
    serii = [s.strip().upper() for s in serii_text.split(',')]
    if seria_mea in serii: return True
    if grupa_cautata in serii: return True
    if (grupa_cautata + "A") in serii or (grupa_cautata + "B") in serii: return True
    return False

# ==========================================
# 3. CITIRE DATE DIN EXCEL (EXTRACT & TRANSFORM)
# ==========================================
def citeste_sheet(fisier, nume_sheet):
    return pd.read_excel(fisier, sheet_name=nume_sheet, header=None)

def detecteaza_coloane_grupa(df, grupa_cautata):
    # Cauta capetele de tabel cu numele grupei/semigrupei (ex: 411Ca si 411Cb)
    sg_a = (grupa_cautata + "a").lower()
    sg_b = (grupa_cautata + "b").lower()
    col_sga, col_sgb = None, None

    for r in range(min(8, len(df))): # Se uita doar pe primele 8 randuri, in header
        for c in df.columns:
            val = normalize_text(df.iloc[r, c]).lower()
            if val == sg_a: col_sga = c
            elif val == sg_b: col_sgb = c

    return col_sga, col_sgb

def grupa_are_semigrupe(fisier, nume_sheet, grupa):
    # Verifica simplu daca pe coloane s-a gasit semigrupa b (pentru a cere input de la user la inceput)
    df = citeste_sheet(fisier, nume_sheet)
    col_sga, col_sgb = detecteaza_coloane_grupa(df, grupa)
    return col_sgb is not None

def detecteaza_coloane_anunturi(df):
    # Scanare Verticala pt pagina de anunturi: citeste header-ul ca sa stie pe ce coloana e Sala, Ora, etc.
    col_disc, col_serii, col_sala, col_zi_ora = 0, 1, 2, -1
    for r in range(min(20, len(df))):
        gasit = False
        for c in range(len(df.columns)):
            val = str(df.iloc[r, c]).strip().lower()
            if val in ["sala", "sala / room"]:
                col_sala = c
                gasit = True
            elif "disciplin" in val or "proiecte" in val:
                col_disc = c
            elif "serii" in val or "grup" in val:
                col_serii = c
            elif "ziua" in val or ("ora" in val and "laborat" not in val):
                col_zi_ora = c
        if gasit: break # S-a oprit cand a gasit header-ul
        
    # Daca secretariatul a uitat sa puna titlu pentru Ziua/Ora, deducem ca e in stanga coloanei de Sală
    if col_zi_ora == -1:
        col_zi_ora = col_sala - 1 if col_sala > 0 else 2
    return col_disc, col_serii, col_sala, col_zi_ora

def extrage_sali_laborator(fisier_curatat, nume_sheet_anunturi, grupa_cautata, traducator_activ):
    dataframe = citeste_sheet(fisier_curatat, nume_sheet_anunturi)
    col_disc, col_serii, col_sala, _ = detecteaza_coloane_anunturi(dataframe)
    
    # Trage numele disciplinelor in jos peste celulele lasate goale din lene de secretariat
    dataframe[col_disc] = dataframe[col_disc].ffill()
    dictionar_sali = {}

    for rand in range(len(dataframe)):
        row_vals = [str(x).strip().lower() for x in dataframe.iloc[rand]]
        # Daca vede un alt cap de tabel inserat in mijloc, isi reseteaza coloanele
        if any("disciplin" in x for x in row_vals):
            for c, val in enumerate(row_vals):
                if "disciplin" in val or "proiecte" in val: col_disc = c
                elif "serii" in val or "grup" in val: col_serii = c
                elif val in ["sala", "sala / room"]: col_sala = c

        raw_col0 = str(dataframe.iloc[rand, col_disc]).strip()
        # Filtreaza gunoaiele si titlurile intermediare
        if raw_col0.lower() in ['nan', '', 'disciplina'] or raw_col0.lower().startswith('sali') or raw_col0.lower().startswith('proiecte'):
            continue
        if "optionale" in raw_col0.lower(): break # Opreste aici, optionalele le facem in alta functie

        disciplina_bruta = normalize_key(raw_col0)
        seriile = normalize_text(dataframe.iloc[rand, col_serii]) if col_serii < len(dataframe.columns) else ""
        sala = normalize_text(dataframe.iloc[rand, col_sala]) if col_sala < len(dataframe.columns) else ""

        if sala.lower() in ['nan', '', '-', '--']: continue
        if not seria_se_potriveste(seriile, grupa_cautata): continue # Ignora materiile de la alte serii

        dictionar_sali[disciplina_bruta] = sala

        # Traducem numele lung din Anunturi in cel scurt din Orar
        for prescurtare, nume_lung in traducator_activ.items():
            p_norm = normalize_key(prescurtare)
            n_norm = normalize_key(nume_lung)

            # Evitam potriviri periculoase pe randuri mixte de tip "Fizica / Sport"
            if '/' in disciplina_bruta:
                continue
            # Salvam si numele scurt si numele lung
            if disciplina_bruta == n_norm or disciplina_bruta == p_norm:
                dictionar_sali[p_norm] = sala
                dictionar_sali[n_norm] = sala

    return dictionar_sali

def extrage_optiuni_disponibile(fisier_curatat, nume_sheet_anunturi, grupa_cautata):
    # Citeste optiunile din Excel si construieste un meniu numerotat de tip [1, 2, 3...]
    dataframe = citeste_sheet(fisier_curatat, nume_sheet_anunturi)
    col_disc, col_serii, col_sala, col_zi_ora = detecteaza_coloane_anunturi(dataframe)
    dataframe[col_disc] = dataframe[col_disc].ffill()
    
    optiuni = {}
    tip_curent = None

    for rand in range(len(dataframe)):
        row_vals = [str(x).strip().lower() for x in dataframe.iloc[rand]]
        if any("disciplin" in x for x in row_vals):
            for c, val in enumerate(row_vals):
                if "disciplin" in val: col_disc = c
                elif "serii" in val or "grup" in val: col_serii = c
                elif "ziua" in val or ("ora" in val and "laborat" not in val): col_zi_ora = c
                elif val in ["sala", "sala / room"]: col_sala = c

        col0 = normalize_text(dataframe.iloc[rand, col_disc])
        col0_lower = col0.lower()

        # Cand intram in Opționale / Facultative, activeaza tipul curent
        if "optional" in col0_lower or "opțional" in col0_lower:
            tip_curent = "Optional"
            continue
        elif "facultativ" in col0_lower or "liber" in col0_lower:
            tip_curent = "Facultativ"
            continue

        if tip_curent is None or col0 in ['nan', 'Disciplina']: continue

        # Uneste variantele Ro/En cu un slash (ex: "Sociologie / Sociology")
        linii_disc = [x.strip() for x in col0.split('\n') if x.strip()]
        disciplina = " / ".join(linii_disc) if len(linii_disc) > 1 else col0.strip()

        seriile = normalize_text(dataframe.iloc[rand, col_serii]) if col_serii < len(dataframe.columns) else ""
        raw_ziua = str(dataframe.iloc[rand, col_zi_ora]).strip().replace(',', ' / ') if col_zi_ora < len(dataframe.columns) else ""
        raw_sala = str(dataframe.iloc[rand, col_sala]).strip().replace('\n', ' / ') if col_sala < len(dataframe.columns) else ""

        if raw_ziua == 'nan' or raw_sala == 'nan' or not raw_ziua: continue

        # Curatam modul in care sunt scrise orele, pastrand asocierile corecte cu ziua
        zile_saptamana = ['LUNI', 'MARTI', 'MIERCURI', 'JOI', 'VINERI']
        linii_zo = [x.strip() for x in raw_ziua.split('\n') if x.strip()]
        curat_zo = []
        for linie in linii_zo:
            are_zi = any(z in linie.upper() for z in zile_saptamana)
            if are_zi or not curat_zo:
                curat_zo.append(linie)
            else:
                curat_zo[-1] += f" / {linie}"

        ziua_ora = " / ".join(curat_zo)
        ziua_ora = re.sub(r'\s*/\s*/\s*', ' / ', ziua_ora)

        # Adauga in meniu (Dictionar) doar materiile valabile pentru grupa ta, si le asigneaza un ID crescator (1, 2...)
        if seria_se_potriveste(seriile, grupa_cautata):
            optiuni[len(optiuni) + 1] = {
                "Nume": disciplina,
                "Zi_Ora": ziua_ora,
                "Sala": raw_sala,
                "Tip": tip_curent
            }

    return optiuni

def detecteaza_coloana_ora(df, col_sga):
    # Pleaca de la coloana cu grupa ta si merge cu -1 spre stanga pentru a gasi Timpul
    if col_sga is None: return 1
    for c in range(col_sga - 1, 0, -1):
        ora_count = 0
        for r in range(2, min(30, len(df))):
            if extrage_ore_slots(df.iloc[r, c]): ora_count += 1
        # Daca a numarat macar 3 randuri cu format logic de ora (ex 08-10), stie clar ca e coloana corecta
        if ora_count >= 3: return c
    return 1

def detecteaza_coloana_sala(df, col_sga, grupa_cautata):
    # Pleaca de la coloana ta si se uita cu +1 spre dreapta dupa Sală
    if col_sga is None: return None
    seria_mea = grupa_cautata[-1].lower()
    
    for c in range(col_sga + 1, len(df.columns)):
        alta_serie = False
        for r in range(min(8, len(df))):
            val = normalize_text(df.iloc[r, c]).lower()
            if val in ["sala", "sala / room", "săli"]: return c # Cauta normal dupa titlu
            
            # Daca da de grupa altei serii (ex: noi suntem F si vedem 431Ga), opreste scanarea spre dreapta
            if len(val) >= 4 and val[:2].isdigit() and val[3].isalpha():
                if val[3] != seria_mea: alta_serie = True
                
        if not alta_serie: # Cazul cand secretariatul uita sa scrie titlul Sala
            room_count = 0
            for r in range(8, min(40, len(df))):
                cell_val = normalize_text(df.iloc[r, c])
                # Verifica printr-un filtru regex daca celula arata a sala (ex: Litera + Cifre = "B125", "A05", "BN030")
                if re.match(r'^[A-Z]{1,3}\s*\d{2,4}[A-Z]?$', cell_val.strip(), re.IGNORECASE):
                    room_count += 1
            if room_count >= 2: return c # Daca gaseste minim 2, nu este o coincidenta
            
        if alta_serie: break # Oprim cautarea stricta

    # Doar daca 2 serii impart sala, trece de granita seriei si ia prima coloana Sală disponibila.
    for c in range(col_sga + 1, len(df.columns)):
        for r in range(min(8, len(df))):
            val = normalize_text(df.iloc[r, c]).lower()
            if val in ["sala", "sala / room", "săli"]: return c
            
    return None

def construieste_grila_goala(exista_sg2):
    # Deseneaza scheletul de ore (Luni 09-10, 10-11, etc.) complet liber
    grila = []
    for zi in ZILE_SAPTAMANA:
        for ora in ORE_STANDARD:
            grila.append({
                'Ziua': zi, 'Ora': ora,
                'SG1': {'Impara': 'Liber', 'Para': 'Liber'},
                'SG2': {'Impara': 'Liber' if exista_sg2 else '-', 'Para': 'Liber' if exista_sg2 else '-'},
                'SalaExcel': ''
            })
    return grila

def incarca_date_in_grila(df, grila, col_ora, col_sga, col_sgb, col_sala):
    # Toarna informatia gasita pe coloanele Excelului in grila goala
    index_grila = {(e['Ziua'], e['Ora']): e for e in grila}

    for rand in range(2, len(df)):
        ziua = normalize_text(df.iloc[rand, 0]).upper()
        if ziua not in ZILE_SAPTAMANA: continue

        ore_gasite = extrage_ore_slots(df.iloc[rand, col_ora])
        if not ore_gasite: continue

        # Transforma textul cu materii imparte/pare in dictionar structurat
        mat_a = parseaza_text_materie(normalize_text(df.iloc[rand, col_sga])) if col_sga is not None else {"Impara": "Liber", "Para": "Liber"}
        mat_b = parseaza_text_materie(normalize_text(df.iloc[rand, col_sgb])) if col_sgb is not None else {"Impara": "-", "Para": "-"}

        sala_excel = ""
        if col_sala is not None:
            val_s = normalize_text(df.iloc[rand, col_sala])
            if val_s.lower() not in ['nan', '', '-', '--']: sala_excel = val_s

        for ora_slot in ore_gasite:
            cheia = (ziua, ora_slot)
            if cheia not in index_grila: continue
            slot = index_grila[cheia]

            # Ataseaza materia doar daca slotul e liber
            if mat_a['Impara'] not in ["Liber", "-", ""]: slot['SG1']['Impara'] = mat_a['Impara']
            if mat_a['Para'] not in ["Liber", "-", ""]: slot['SG1']['Para'] = mat_a['Para']

            if col_sgb is not None:
                if mat_b['Impara'] not in ["Liber", "-", ""]: slot['SG2']['Impara'] = mat_b['Impara']
                if mat_b['Para'] not in ["Liber", "-", ""]: slot['SG2']['Para'] = mat_b['Para']

            if sala_excel: slot['SalaExcel'] = sala_excel

def ataseaza_sala_din_dictionar(text_materie, dictionar_sali, traducator_activ, allow_broad=True):
    # Compara o materie cu dictionarul de sali si, daca gaseste potrivire, lipeste sala cu paranteze "[A05]"
    if text_materie == "Liber": return text_materie
    if materia_are_sala(text_materie): return text_materie

    text_norm = normalize_key(text_materie)
    if text_norm in dictionar_sali:
        return f"{text_materie} [{dictionar_sali[text_norm]}]"

    # allow_broad=True permite potriviri partile (utile pentru laboratoare). La curs, o oprim.
    if allow_broad:
        for prescurtare, nume_lung in traducator_activ.items():
            p_norm = normalize_key(prescurtare)
            n_norm = normalize_key(nume_lung)

            if p_norm in text_norm or text_norm == n_norm:
                if p_norm in dictionar_sali: return f"{text_materie} [{dictionar_sali[p_norm]}]"
                if n_norm in dictionar_sali: return f"{text_materie} [{dictionar_sali[n_norm]}]"

    return text_materie

def ataseaza_salile_finale(grila, dictionar_sali, traducator_activ):
    # Se plimba prin toata grila si aplica functia de mai sus pentru fiecare materie
    for e in grila:
        sala_excel = e['SalaExcel']
        for sg in ['SG1', 'SG2']:
            for par in ['Impara', 'Para']:
                mat = e[sg][par]

                if not este_materie(mat):
                    if str(mat).strip() != "Liber": e[sg][par] = "Liber"
                    continue

                if este_activitate_sport(mat) and not materia_are_sala(mat):
                    if scoate_sala_din_text(mat).lower().strip() in {
                        "sport",
                        "sport (s)",
                        "sp(s)",
                        "sp (s)",
                        "educatie fizica",
                        "educaţie fizică",
                        "educatie fizica si sport 2",
                        "educaţie fizică şi sport 2"
                    }:
                        e[sg][par] = f"{mat} [Sala Sport LEU]"
                        continue

                if materia_are_sala(mat): continue

                if este_curs(mat) and not este_activitate_sport(mat):
                    mat_cu_sala = ataseaza_sala_din_dictionar(mat, dictionar_sali, traducator_activ, allow_broad=False)
                else:
                    mat_cu_sala = ataseaza_sala_din_dictionar(mat, dictionar_sali, traducator_activ, allow_broad=True)

                if mat_cu_sala != mat:
                    e[sg][par] = mat_cu_sala
                    continue

                # Daca nu a reusit sa aduca din Dictionarul de Anunturi, ataseaza sala gasita in detecteaza_coloana_sala
                if sala_excel:
                    if not este_activitate_sport(mat) and sala_de_sport(sala_excel): continue
                    e[sg][par] = f"{mat} [{sala_excel}]"

def normalizeaza_sali_cursuri_pe_zi(grila):
    # Daca sala e scrisa doar la semigrupa A (sau la alta grupa), el ia "Sala Dominanta" de pe rand si i-o da si semigrupei B.
    for zi in ZILE_SAPTAMANA:
        intrari_zi = [e for e in grila if e['Ziua'] == zi]
        for sg in ['SG1', 'SG2']:
            for par in ['Impara', 'Para']:
                sali_curs_valide = []
                for e in intrari_zi:
                    mat = e[sg][par]
                    if not este_materie(mat) or not este_curs(mat) or este_activitate_sport(mat): continue
                    sala = extrage_sala_din_text(mat)
                    if sala and not sala_de_sport(sala): sali_curs_valide.append(sala)

                if not sali_curs_valide: continue
                # Decide care e Sala Dominanta numarand cu Counter (cea mai frecventa sala pe orizontală)
                sala_dominanta = Counter(sali_curs_valide).most_common(1)[0][0]

                # Imparte sala la semigrupele care o au lipsa
                for e in intrari_zi:
                    mat = e[sg][par]
                    if not este_materie(mat) or not este_curs(mat) or este_activitate_sport(mat): continue
                    sala = extrage_sala_din_text(mat)
                    if not sala or sala_de_sport(sala):
                        e[sg][par] = inlocuieste_sala_in_text(mat, sala_dominanta)

def genereaza_orar_grupa(fisier_curatat, nume_sheet, grupa_cautata, dictionar_sali, traducator_activ):
    # Centralizeaza tot fluxul de creare a orarului pentru grupa curenta
    df = citeste_sheet(fisier_curatat, nume_sheet)
    df[0] = df[0].ffill()

    col_sga, col_sgb = detecteaza_coloane_grupa(df, grupa_cautata)
    if col_sga is None and col_sgb is None: return []

    col_ora = detecteaza_coloana_ora(df, col_sga)
    col_sala = detecteaza_coloana_sala(df, col_sga, grupa_cautata)

    grila = construieste_grila_goala(exista_sg2=(col_sgb is not None))
    incarca_date_in_grila(df, grila, col_ora, col_sga, col_sgb, col_sala)
    ataseaza_salile_finale(grila, dictionar_sali, traducator_activ)
    normalizeaza_sali_cursuri_pe_zi(grila)

    return grila

# ==========================================
# 4. EXPORT EXCEL (LOAD)
# ==========================================
def determinare_culoare(text, culori):
    # Pe baza unor cuvinte cheie in denumire (ex: "curs", "lab", "sport"), alege ce culoare de fundal primeste celula in Excel.
    text_original = str(text)
    t = scoate_sala_din_text(text_original).lower().strip()

    if "liber" in t or t == "-" or re.match(r'^-+$', t):
        return culori["liber"]

    if "⭐" in text_original or "opt " in t or "optional" in t or re.search(r'(^|\s)o[1-9][\s\.\(]', t):
        return culori["optional"]

    if "sport" in t or "educaţie fizică" in t or "educatie fizica" in t or "sp(s)" in t:
        return culori["seminar"]

    if "(curs)" in t or "(c)" in t or " curs" in t or "(course)" in t or "(lecture)" in t:
        return culori["curs"]

    if "(laborator)" in t or "(l)" in t or " lab" in t:
        return culori["laborator"]

    if "(seminar)" in t or "(s)" in t or " sem" in t:
        return culori["seminar"]

    if "(p)" in t or "proiect" in t or "pr " in t:
        return culori["proiect"]

    if not t or t == "nan":
        return None

    # Materiile necatalogate (fara etichete de "(curs)", "(lab)") primesc automat ROZ
    return culori["optional"]

def stiluri_excel():
    # Returneaza pachetul complet de setari grafice pentru OpenPyxl (Borders, Fonts, Colors)
    culori = {
        "header": PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
        "curs": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
        "seminar": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "laborator": PatternFill(start_color="FF8080", end_color="FF8080", fill_type="solid"),
        "proiect": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "optional": PatternFill(start_color="FF99CC", end_color="FF99CC", fill_type="solid"),
        "liber": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    }
    stil = {
        "culori": culori,
        "font_header": Font(bold=True, color="FFFFFF", name="Calibri", size=12),
        "font_normal": Font(name="Calibri", size=11),
        "font_bold": Font(name="Calibri", size=11, bold=True),
        "aliniere": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border_subtire": Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    }
    return stil

def aplica_stil_celula(cell, font, alignment, border, fill=None):
    # Aplicator scurt pt ca reduce numarul de linii de cod din loop-ul de desenare
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill: cell.fill = fill

def adauga_optionale_peste_orar(mat_sg1, mat_sg2, optionale_alese, zi, ora_start, paritate, semigrupa_utilizator, are_sg2):
    # Ia optiunile userului si suprascrie in mod tintit orarul general.
    for opt in optionale_alese:
        # Verifica daca se nimeresc la zi, ora si saptamana (impara/para)
        if opt['Ziua'] == zi and opt['OraStart'] <= ora_start < opt['OraStop']:
            if opt['Paritate'] in ["Ambele", paritate]:
                # Modifica DOAR casuta semigrupei alese, pentru a nu strica orarul de langa el
                if semigrupa_utilizator == 'a': mat_sg1 = f"⭐ {opt['Nume']}"
                elif semigrupa_utilizator == 'b' and are_sg2: mat_sg2 = f"⭐ {opt['Nume']}"
    return mat_sg1, mat_sg2

def uneste_intervale_identice(ws, start_row, end_row, are_sg2):
    # Citeste blocuri repetate de 1 ora (10-11 Fizica, 11-12 Fizica) si le transforma prin Merge Cells intr-un bloc unitar (10-12 Fizica).
    def baza(val): return scoate_sala_din_text(str(val))
    def get_sala(val): return extrage_sala_din_text(str(val))

    # Mentine in memorie o copie neafectata inainte de lipire
    valori_orig = {}
    for row in range(start_row, end_row + 1):
        valori_orig[row] = {
            'ora': str(ws.cell(row=row, column=2).value).strip(),
            'sg1': baza(ws.cell(row=row, column=3).value),
            'sg2': baza(ws.cell(row=row, column=4).value) if are_sg2 else None
        }

    # Lipește Coloana Timpului. Calculeaza ora finala din interval si le imbina DOAR daca Semigrupele fac acelasi lucru
    r = start_row
    while r <= end_row:
        b1 = valori_orig[r]['sg1']
        b2 = valori_orig[r]['sg2']
        r_end = r
        while r_end + 1 <= end_row:
            if valori_orig[r_end + 1]['sg1'] == b1 and (not are_sg2 or valori_orig[r_end + 1]['sg2'] == b2):
                r_end += 1
            else:
                break
        
        if r_end > r:
            ora_start = valori_orig[r]['ora'].split('-')[0]
            ora_stop = valori_orig[r_end]['ora'].split('-')[1]
            for i in range(r + 1, r_end + 1):
                ws.cell(row=i, column=2).value = ""
            ws.cell(row=r, column=2).value = f" {ora_start}-{ora_stop}" # Ex: 10-12
            ws.merge_cells(start_row=r, start_column=2, end_row=r_end, end_column=2)
        r = r_end + 1

    # Lipește coloanele Semigrupelor in mod separat
    coloane = [3, 4] if are_sg2 else [3]
    for col in coloane:
        col_key = 'sg1' if col == 3 else 'sg2'
        r = start_row
        while r <= end_row:
            b = valori_orig[r][col_key]
            r_end = r
            while r_end + 1 <= end_row:
                if valori_orig[r_end + 1][col_key] == b:
                    r_end += 1
                else:
                    break
            
            if r_end > r:
                valori_celule = [ws.cell(row=i, column=col).value for i in range(r, r_end + 1)]
                text_final = next((v for v in valori_celule if get_sala(v)), valori_celule[0])
                
                for i in range(r + 1, r_end + 1):
                    ws.cell(row=i, column=col).value = ""
                
                ws.cell(row=r, column=col).value = text_final
                ws.merge_cells(start_row=r, start_column=col, end_row=r_end, end_column=col)
            r = r_end + 1

def traseaza_linie_final_zi(ws, end_row, nr_coloane):
    # La final de zi (Luni, Marti) pune o bordura mai groasa la baza
    for col_idx in range(1, nr_coloane + 1):
        cell_baza = ws.cell(row=end_row, column=col_idx)
        margine_veche = cell_baza.border
        cell_baza.border = Border(left=margine_veche.left, right=margine_veche.right, top=margine_veche.top, bottom=Side(style='medium'))

def creeaza_sheet_paritate(wb, orar_baza, optionale_alese, grupa, semigrupa_utilizator, are_sg2, nume_sheet, paritate, stil):
    # Creaza o pagina (Saptamana Impara sau Para) si scrie randurile de orar in fisierul final OpenPyxl
    ws = wb.create_sheet(title=nume_sheet)
    headers = ["Ziua", "Ora", f"SG1 ({grupa}a)"]
    if are_sg2: headers.append(f"SG2 ({grupa}b)")
    ws.append(headers)

    for cell in ws[1]:
        aplica_stil_celula(cell, stil["font_header"], stil["aliniere"], stil["border_subtire"], stil["culori"]["header"])

    rand_curent = 2
    for zi in ZILE_SAPTAMANA:
        intrari_zi = [e for e in orar_baza if e['Ziua'] == zi]
        if not intrari_zi: continue

        start_row = rand_curent
        for intrare in intrari_zi:
            ora = intrare['Ora']
            ora_start = int(ora.split('-')[0])

            mat_sg1 = intrare['SG1'][paritate]
            mat_sg2 = intrare['SG2'][paritate] if are_sg2 else "-"

            # Momentul in care bagam Opționalele din consola peste grila normala (inainte sa coloram)
            mat_sg1, mat_sg2 = adauga_optionale_peste_orar(mat_sg1, mat_sg2, optionale_alese, zi, ora_start, paritate, semigrupa_utilizator, are_sg2)

            rand_nou = [zi, f" {ora}", mat_sg1]
            if are_sg2: rand_nou.append(mat_sg2)
            ws.append(rand_nou)

            # Aplica stilurile + Vopseaua potrivita
            nr_coloane = 4 if are_sg2 else 3
            for col_idx in range(1, nr_coloane + 1):
                cell = ws.cell(row=rand_curent, column=col_idx)
                aplica_stil_celula(cell, stil["font_bold"] if col_idx == 1 else stil["font_normal"], stil["aliniere"], stil["border_subtire"])

                if col_idx == 3:
                    fill = determinare_culoare(str(mat_sg1), stil["culori"])
                    if fill: cell.fill = fill
                elif col_idx == 4 and are_sg2:
                    fill = determinare_culoare(str(mat_sg2), stil["culori"])
                    if fill: cell.fill = fill

            rand_curent += 1

        end_row = rand_curent - 1
        uneste_intervale_identice(ws, start_row, end_row, are_sg2) # Lipim orele duble
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1) # Lipeste coloana de 'Luni' pe verticala

        nr_coloane = 4 if are_sg2 else 3
        traseaza_linie_final_zi(ws, end_row, nr_coloane)

    # Setam marimile coloanelor
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 45
    if are_sg2: ws.column_dimensions['D'].width = 45

def exporteaza_in_excel(orar_baza, optionale_alese, grupa, semigrupa_utilizator):
    # Sterge foaia goala initiala a Excelului, da start generarii paginilor, apoi salveaza!
    if not orar_baza: return

    are_sg2 = False
    for intrare in orar_baza:
        for par in ['Impara', 'Para']:
            if este_materie(intrare['SG2'][par]):
                are_sg2 = True
                break
        if are_sg2: break

    wb = Workbook()
    wb.remove(wb.active)
    stil = stiluri_excel()

    creeaza_sheet_paritate(wb, orar_baza, optionale_alese, grupa, semigrupa_utilizator, are_sg2, "Saptamana_Impara", "Impara", stil)
    creeaza_sheet_paritate(wb, orar_baza, optionale_alese, grupa, semigrupa_utilizator, are_sg2, "Saptamana_Para", "Para", stil)

    nume_fisier_iesire = f"Orar_Personalizat_{grupa}{semigrupa_utilizator}.xlsx"
    wb.save(nume_fisier_iesire)
    print(f"\nOrarul a fost generat si salvat ca '{nume_fisier_iesire}'!")

# ==========================================
# 5. MENIU OPTIONALE
# ==========================================
def parseaza_timp_optional(timp_ales):
    # Traduce textul salbatic "vineri impar 17- 19" in date procesabile: {'Ziua': 'VINERI', 'OraStart': 17, 'OraStop': 19, 'Paritate': 'Impara'}
    timp_upper = timp_ales.upper()
    zile_saptamana = ['LUNI', 'MARTI', 'MIERCURI', 'JOI', 'VINERI']

    ziua_aleasa = next((z for z in zile_saptamana if z in timp_upper), None)
    if not ziua_aleasa: return None

    timp_lower = timp_ales.lower()
    paritate = "Ambele"
    if "impar" in timp_lower: paritate = "Impara"
    elif "par" in timp_lower: paritate = "Para"

    match = re.search(r'(\d{1,2})\s*-\s*(\d{1,2})', timp_ales)
    if not match: return None

    return {'Ziua': ziua_aleasa, 'OraStart': int(match.group(1)), 'OraStop': int(match.group(2)), 'Paritate': paritate}

def proceseaza_alegeri_meniu(alegeri_str, dict_optiuni, optionale_alese_final):
    # Parcurge lista de numere introduse in consola (ex: "1,3") si gaseste detaliile materiilor in dictionar
    numere_alese = alegeri_str.split(',')

    for numar in numere_alese:
        numar = numar.strip()
        if not (numar.isdigit() and int(numar) in dict_optiuni): continue

        opt = dict_optiuni[int(numar)]
        nume_mat = opt['Nume']
        
        raw_ziua = opt['Zi_Ora']
        raw_sala = opt['Sala']
        
        # Curata si izoleaza intervalele (in caz ca o materie are mai multe variante de orar)
        raw_ziua_clean = re.sub(r'(?i)s[aă]pt[a-z\.\s]*\d{1,2}\s*-\s*\d{1,2}', '', raw_ziua)
        zile_saptamana = ['LUNI', 'MARTI', 'MIERCURI', 'JOI', 'VINERI']
        ziua_gasita = next((z for z in zile_saptamana if z in raw_ziua_clean.upper()), "")

        matches = re.findall(r'(\d{1,2})\s*-\s*(\d{1,2})', raw_ziua_clean)
        if not matches: continue
        
        variante_timp = [f"{ziua_gasita} {m[0]}-{m[1]}" for m in matches]
        salile = [s.strip() for s in raw_sala.replace('\n', '/').split('/') if s.strip()]

        optiuni_prezentate = []
        for i, v_timp in enumerate(variante_timp):
            s_val = salile[i] if i < len(salile) else (salile[-1] if salile else "")
            optiuni_prezentate.append((v_timp, s_val))

        # Sterge duplicatele posibile din excel
        optiuni_prezentate = list(dict.fromkeys(optiuni_prezentate))

        # Daca exista doar o varinata de orar pt Optiunea asta, o alege direct
        if len(optiuni_prezentate) == 1:
            t_ales, s_aleasa = optiuni_prezentate[0]
            sala_disp = f" [{s_aleasa}]" if s_aleasa else ""
            print(f"\nS-a adaugat automat: {nume_mat} -> {t_ales}{sala_disp}")
            timp_ales, sala_aleasa = t_ales, s_aleasa
        else:
            # Daca materia se tine in 2 grupe la ore diferite, primeste alt input
            print(f"\nPentru '{nume_mat}', ai {len(optiuni_prezentate)} variante:")
            for i, (t_val, s_val) in enumerate(optiuni_prezentate):
                s_disp = s_val if s_val else "N/A"
                print(f"  [{i+1}] {t_val} | Sala: {s_disp}")

            alegere_v = input(f"Alege varianta dorita (1-{len(optiuni_prezentate)}): ").strip()
            if not (alegere_v.isdigit() and 1 <= int(alegere_v) <= len(optiuni_prezentate)):
                print("Alegere invalida. Sarim peste.")
                continue

            idx = int(alegere_v) - 1
            timp_ales, sala_aleasa = optiuni_prezentate[idx]

        # Formateaza corect item-ul pt dictionar
        timp_parsat = parseaza_timp_optional(timp_ales)
        if not timp_parsat:
            print(f"Nu am putut interpreta ora din '{timp_ales}'")
            continue

        sala_text_final = f" [{sala_aleasa}]" if sala_aleasa else ""
        optionale_alese_final.append({
            'Nume': f"{nume_mat}{sala_text_final}",
            **timp_parsat
        })

# ==========================================
# 6. PROGRAM PRINCIPAL
# ==========================================
def main():
    fisier = 'orar_curatat.xlsx'

    grupa, anul = extrage_grupa_si_anul()

    # Seria G mereu engleza, Seria F engleza doar din anul 3
    este_engleza = grupa.endswith('G') or (grupa.endswith('F') and anul in ['3', '4'])
    traducator_activ = TRADUCATOR_EN if este_engleza else TRADUCATOR_RO

    pagina_an = f"AN {anul}"
    pagina_anunturi = f"Anunturi - AN{anul}"

    try:
        are_semigrupe = grupa_are_semigrupe(fisier, pagina_an, grupa)
    except Exception as e:
        print(f"\nEroare la citirea paginii '{pagina_an}'.")
        print(f"   Detaliu tehnic: {e}")
        return

    # Daca exista semigrupe, il punem sa isi aleaga a sau b
    if are_semigrupe:
        semigrupa_utilizator = input("Te rog introdu semigrupa ta ('a' sau 'b'): ").strip().lower()
        if semigrupa_utilizator not in ['a', 'b']:
            print("Semigrupa invalida. Vom folosi default 'a'.")
            semigrupa_utilizator = 'a'
    else:
        semigrupa_utilizator = 'a'

    # PORNESTE FLUXUL DE PROCESARE
    try:
        sali_extrase = extrage_sali_laborator(fisier, pagina_anunturi, grupa, traducator_activ)
        meniu_optiuni = extrage_optiuni_disponibile(fisier, pagina_anunturi, grupa)
        orar_baza = genereaza_orar_grupa(fisier, pagina_an, grupa, sali_extrase, traducator_activ)
    except Exception as e:
        print(f"\nEroare la citirea paginilor pentru Anul {anul}.")
        print(f"   Verifica daca sheet-urile se numesc exact '{pagina_an}' si '{pagina_anunturi}'.")
        print(f"   Detaliu tehnic: {e}")
        return

    optionale_alese_final = []

    # MENIURI OPTIONALE
    dict_optionale = {k: v for k, v in meniu_optiuni.items() if v["Tip"] == "Optional"}
    if dict_optionale:
        print("Acestea sunt optionalele disponibile pentru seria ta:")
        for id_opt, detalii in dict_optionale.items():
            print(f" [{id_opt}] {detalii['Nume']} | {detalii['Zi_Ora'].replace('/','').replace('  ', ' ')}")

        alegeri_opt = input("\nScrie numerele materiilor alese (ex: 1,3) sau Enter daca nu vrei: ")
        if alegeri_opt.strip():
            proceseaza_alegeri_meniu(alegeri_opt, dict_optionale, optionale_alese_final)

    # MENIURI FACULTATIVE
    dict_facultative = {k: v for k, v in meniu_optiuni.items() if v["Tip"] == "Facultativ"}
    if dict_facultative:
        rasp_fac = input("Doresti sa adaugi materii facultative? (Da/Nu): ").strip().lower()

        if rasp_fac == 'da':
            for id_opt, detalii in dict_facultative.items():
                print(f" [{id_opt}] {detalii['Nume']} | {detalii['Zi_Ora'].replace('/','').replace('  ', ' ')}")

            alegeri_fac = input("\nScrie numerele alese: ")
            if alegeri_fac.strip():
                proceseaza_alegeri_meniu(alegeri_fac, dict_facultative, optionale_alese_final)

    if not orar_baza:
        print("\nNu s-a gasit orar pentru grupa specificata.")
        return

    # DATELE MERG SPRE EXCEL PENTRU EXPORTUL FINAL
    exporteaza_in_excel(orar_baza, optionale_alese_final, grupa, semigrupa_utilizator)

if __name__ == "__main__":
    main()