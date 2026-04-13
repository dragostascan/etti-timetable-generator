import openpyxl

def curata_excel_orar(fisier_intrare, fisier_iesire):
    workbench = openpyxl.load_workbook(fisier_intrare)
    
    # Extragem lista cu toate numele paginilor
    toate_paginile = workbench.sheetnames
    
    # Luam doar primele 4 pagini folosind slicing
    pagini_orar = toate_paginile[:4]

    # Acum parcurgem fiecare pagina pe rand
    for nume_pagina in pagini_orar:
        sheet = workbench[nume_pagina]
        
        merged_ranges = list(sheet.merged_cells.ranges)

        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            
            sheet.unmerge_cells(str(merged_range))
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col).value = top_left_cell_value

    print(f"Salvez varianta curatata in {fisier_iesire}")
    workbench.save(fisier_iesire)
    return fisier_iesire

curata_excel_orar('orar_etti.xlsx', 'orar_curatat.xlsx')